import time

import pandas as pd
import requests
import shutil
import os
import sqlite3
import ast

from openpyxl import load_workbook


def query_products(con, link):
    cursor = con.cursor()
    cmd = cursor.execute(f"SELECT * FROM products WHERE link = '{link}'")
    data = cmd.fetchone()
    return data

def save_data(data):
    sheet_name = f"products"
    try:
        if not os.path.isfile("products.xlsx"):
            dataframe = pd.DataFrame(data)
            data_to_excel = pd.ExcelWriter('products.xlsx', engine='xlsxwriter')
            dataframe.to_excel(data_to_excel, sheet_name=sheet_name,index=False)
            data_to_excel.save()
            # data_to_excel.close()
        else:
            wb = load_workbook("products.xlsx", read_only=True)

            if str(sheet_name) in wb.sheetnames:
                file = pd.read_excel('products.xlsx',sheet_name=str(sheet_name))
                new_df = file.append(data)
                with pd.ExcelWriter('products.xlsx', engine='openpyxl', mode="a", if_sheet_exists="overlay") as writes:
                        new_df.to_excel(writes,str(sheet_name),index=False)
            else:
                dataframe = pd.DataFrame(data)
                with pd.ExcelWriter('products.xlsx', engine='openpyxl', mode="a", if_sheet_exists="overlay") as writes:
                    dataframe.to_excel(writes, str(sheet_name), index=False)
    except:
        pass


def save_image(con, name,link):
    cur = con.cursor()
    cur.execute("INSERT INTO downloads (path, link) VALUES (?,?)",(name, link))
    con.commit()

def download_image(con, product_count, image):
    try:
        r = requests.get(image, stream=True)
        if r.status_code == 200:
            if not os.path.isdir(f"recomendation/{product_count}"):
                os.mkdir(f"recomendation/{product_count}")
                os.mkdir(f"recomendation/{product_count}/Query")
            with open(f"recomendation/{product_count}/Query/" + image.split("/")[-1], 'wb') as f:
                r.raw.decode_content = True
                shutil.copyfileobj(r.raw, f)
            save_image(con, f"recomendation/{product_count}/Query/" + image.split("/")[-1], image)
            return f"recomendation/{product_count}/Query/" + image.split("/")[-1]
    except Exception as e:
        exit(e)

if __name__ == "__main__":
    save_data([{
        "title": "title",
        "price": "5",
        "desc": "fgfdgdf",
    }])